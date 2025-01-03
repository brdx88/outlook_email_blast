import openpyxl
import os
import win32com.client as win32

def create_excel_file(file_path):
    # Create a new Excel file
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sample Sheet"
    ws["A1"] = "Hello"
    ws["B1"] = "World"
    wb.save(file_path)
    print(f"Excel file created at {file_path}")

def encrypt_excel_file(input_file, password):
    # Use pywin32 to set a password on the Excel file
    excel = win32.gencache.EnsureDispatch("Excel.Application")
    excel.Visible = False  # Run Excel in the background

    # Check if the file exists
    if not os.path.exists(input_file):
        print(f"Error: The file '{input_file}' does not exist.")
        return None

    workbook = excel.Workbooks.Open(input_file)
    encrypted_file = input_file.replace(".xlsx", "_protected.xlsx")

    workbook.Password = password  # Set the password
    workbook.SaveAs(encrypted_file, FileFormat=51)  # FileFormat=51 for .xlsx files
    workbook.Close()
    excel.Quit()
    print(f"Encrypted file saved at {encrypted_file}")
    return encrypted_file

# Define paths
input_file = os.path.abspath("sample.xlsx")  # Use absolute path
password = "securepassword123"

# Create Excel file
create_excel_file(input_file)

# Encrypt the file
encrypted_file = encrypt_excel_file(input_file, password)

# Clean up the original file (optional)
if encrypted_file:
    os.remove(input_file)
